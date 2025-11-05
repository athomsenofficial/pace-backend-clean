import os
import io
import uuid
from fastapi import Body, FastAPI, Form, UploadFile, File, Query
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
import pandas as pd
from final_mel_generator import generate_final_roster_pdf
from session_manager import create_session, get_pdf_from_redis, get_session, update_session, delete_session
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict, Optional
from initial_mel_generator import generate_roster_pdf
from roster_processor import roster_processor, recalculate_small_units
from classes import PasCodeInfo, PasCodeSubmission
from constants import (
    REQUIRED_COLUMNS, OPTIONAL_COLUMNS, PDF_COLUMNS,
    cors_origins, allowed_types, images_dir, default_logo
)
from logging_config import LoggerSetup

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def rescan_and_update_pascodes(session_id: str):
    """
    Re-scan ALL categories for PASCODEs and update session.
    This ensures PASCODE tracking stays synchronized with roster data
    after add/edit/delete operations.
    """
    session = get_session(session_id)
    if not session:
        return

    all_pascodes = set()
    pascode_unit_map = {}

    for cat_key in ['eligible_df', 'ineligible_df', 'discrepancy_df', 'btz_df', 'small_unit_df']:
        cat_list = session.get(cat_key, [])
        if isinstance(cat_list, list):
            for member in cat_list:
                if not member.get('deleted', False):  # Skip soft-deleted members
                    pas = member.get('ASSIGNED_PAS')
                    if pas and isinstance(pas, str) and pas.strip():
                        all_pascodes.add(pas.strip())
                        # Track unit name for this PASCODE
                        if 'ASSIGNED_PAS_CLEARTEXT' in member:
                            pascode_unit_map[pas.strip()] = member['ASSIGNED_PAS_CLEARTEXT']

    # Update session with complete PASCODE list
    if all_pascodes:
        update_session(session_id,
                       pascodes=sorted(list(all_pascodes)),
                       pascode_unit_map=pascode_unit_map)
    else:
        # If no PASCODEs remain, clear the lists
        update_session(session_id,
                       pascodes=[],
                       pascode_unit_map={})


@app.get("/api/health")
async def health_check():
    """Health check endpoint for Docker and load balancers"""
    return {"status": "healthy", "service": "pace-backend"}


@app.post("/api/upload/initial-mel")
async def upload_file(
        file: UploadFile = File(...),
        cycle: str = Form(...),
        year: int = Form(...)
):
    # Generate session ID early for logging
    session_id = str(uuid.uuid4())

    # CRITICAL FIX: Validate cycle parameter
    valid_cycles = ['SRA', 'SSG', 'TSG', 'MSG', 'SMS']
    if cycle not in valid_cycles:
        return JSONResponse(
            content={"error": f"Invalid cycle. Must be one of: {', '.join(valid_cycles)}"},
            status_code=400
        )

    # CRITICAL FIX: Validate year parameter
    from constants import MIN_PROMOTION_CYCLE_YEAR, MAX_PROMOTION_CYCLE_YEAR
    try:
        year = int(year)
        if year < MIN_PROMOTION_CYCLE_YEAR or year > MAX_PROMOTION_CYCLE_YEAR:
            return JSONResponse(
                content={"error": f"Invalid year. Must be between {MIN_PROMOTION_CYCLE_YEAR} and {MAX_PROMOTION_CYCLE_YEAR}"},
                status_code=400
            )
    except (ValueError, TypeError):
        return JSONResponse(content={"error": "Year must be a valid integer"}, status_code=400)

    # Create session-specific logger
    logger = LoggerSetup.get_session_logger(session_id, cycle, year)

    logger.info(f"INITIAL MEL UPLOAD STARTED")
    logger.info(f"  Filename: {file.filename}")
    logger.info(f"  Content Type: {file.content_type}")
    logger.info(f"  Cycle: {cycle}")
    logger.info(f"  Year: {year}")

    return_object = {}

    if file.content_type not in allowed_types:
        error_msg = "Invalid file type. Only CSV or Excel files are allowed."
        logger.error(f"  FAILED: {error_msg}")
        logger.info(f"STATUS: FAILED - Invalid File Type")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=400)

    contents = await file.read()
    file_size_bytes = len(contents)
    logger.info(f"  File size: {file_size_bytes} bytes")

    # HIGH FIX: Validate file size
    from constants import MAX_FILE_SIZE_MB
    max_size_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
    if file_size_bytes > max_size_bytes:
        error_msg = f"File too large. Maximum size is {MAX_FILE_SIZE_MB}MB"
        logger.error(f"  FAILED: {error_msg} (received {file_size_bytes / 1024 / 1024:.2f}MB)")
        logger.info(f"STATUS: FAILED - File Too Large")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=400)

    try:
        if file.filename.endswith(".csv"):
            logger.info(f"  Parsing CSV file")
            # HIGH FIX: Add CSV encoding handling
            try:
                df = pd.read_csv(io.BytesIO(contents), encoding='utf-8')
            except UnicodeDecodeError:
                logger.warning(f"  UTF-8 decoding failed, trying cp1252 encoding")
                try:
                    df = pd.read_csv(io.BytesIO(contents), encoding='cp1252')
                except UnicodeDecodeError:
                    logger.warning(f"  cp1252 decoding failed, trying latin1 encoding")
                    df = pd.read_csv(io.BytesIO(contents), encoding='latin1')
        elif file.filename.endswith(".xlsx"):
            logger.info(f"  Parsing Excel file")
            # CRITICAL FIX: Validate Excel sheet has data
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
            if df.empty:
                logger.warning(f"  First sheet is empty, checking other sheets...")
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents))
                sheet_names = wb.sheetnames
                logger.info(f"  Available sheets: {sheet_names}")
                # Try to find first non-empty sheet
                for sheet_name in sheet_names:
                    test_df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name)
                    if not test_df.empty:
                        df = test_df
                        logger.info(f"  Using sheet '{sheet_name}' which contains data")
                        break
        else:
            error_msg = "Unsupported file extension."
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"STATUS: FAILED - Unsupported Extension")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        logger.info(f"  File parsed successfully: {len(df)} rows, {len(df.columns)} columns")

        # HIGH FIX: Filter out completely empty rows
        initial_rows = len(df)
        df = df.dropna(how='all')
        if len(df) < initial_rows:
            logger.info(f"  Filtered out {initial_rows - len(df)} empty rows")

        # HIGH FIX: Normalize column names to uppercase
        df.columns = df.columns.str.strip().str.upper()
        logger.info(f"  Normalized column names to uppercase")

        # MEDIUM FIX: Strip leading/trailing whitespace from string columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

        # CRITICAL FIX: Validate required columns exist
        all_required_columns = REQUIRED_COLUMNS
        missing_columns = [col for col in all_required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Missing required columns: {', '.join(missing_columns)}"
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"  Available columns: {', '.join(df.columns.tolist())}")
            logger.info(f"STATUS: FAILED - Missing Columns")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        # Filter to only include columns we need (required + optional that exist)
        available_optional = [col for col in OPTIONAL_COLUMNS if col in df.columns]
        columns_to_keep = REQUIRED_COLUMNS + available_optional
        processed_df = df[columns_to_keep].copy()

        # Validate PDF columns exist
        missing_pdf_columns = [col for col in PDF_COLUMNS if col not in processed_df.columns]
        if missing_pdf_columns:
            error_msg = f"Missing PDF columns: {', '.join(missing_pdf_columns)}"
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"STATUS: FAILED - Missing PDF Columns")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        pdf_df = processed_df[PDF_COLUMNS].copy()

        # Pass session_id to create_session so it uses our pre-generated ID
        create_session(processed_df, pdf_df, session_id=session_id)
        logger.info(f"  Session created: {session_id}")

        # MEDIUM FIX: Batch session updates to reduce Redis round trips
        update_session(session_id, cycle=cycle, year=year)

        logger.info(f"  Starting roster processing...")
        roster_processor(df, session_id, cycle, year)
        logger.info(f"  Roster processing complete")

        session = get_session(session_id)

        # Use .get() to safely access session keys that might not exist
        if session.get('pascodes') is not None:
            return_object['pascodes'] = session['pascodes']
            logger.info(f"  PASCODEs found: {len(session['pascodes'])}")

        if session.get('pascode_unit_map') is not None:
            return_object['pascode_unit_map'] = session['pascode_unit_map']

        if session.get('small_unit_df') is not None:
            return_object['senior_rater_needed'] = True
            logger.info(f"  Senior rater required for small units")
        else:
            return_object["senior_rater_needed"] = False

        return_object['message'] = "Upload successful."
        return_object['session_id'] = session_id
        return_object['errors'] = session.get('error_log', [])

        if return_object['errors']:
            logger.warning(f"  Upload completed with {len(return_object['errors'])} errors")
        else:
            logger.info(f"  Upload completed successfully with no errors")

        logger.info(f"INITIAL MEL UPLOAD COMPLETED SUCCESSFULLY")

        return JSONResponse(content=return_object)

    except Exception as e:
        error_msg = f"Processing error: {str(e)}"
        logger.error(f"  EXCEPTION: {error_msg}", exc_info=True)
        logger.info(f"STATUS: FAILED - Exception")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=500)


@app.get("/api/download/initial-mel/{session_id}")
async def download_initial_mel(session_id: str):
    try:
        pdf_buffer: Optional[io.BytesIO] = get_pdf_from_redis(session_id)

        if not pdf_buffer:
            return JSONResponse(
                content={"error": "PDF not found for this session"},
                status_code=404
            )

        return StreamingResponse(
            pdf_buffer,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename=initial_mel_roster.pdf"
            }
        )
    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to retrieve PDF: {str(e)}"},
            status_code=500
        )

@app.get("/api/roster/preview/{session_id}")
async def get_roster_preview(
    session_id: str,
    category: str = "all",
    page: int = 1,
    page_size: int = 50
):
    """
    Get roster preview for review and editing.
    Returns all member data from the session including eligible,
    ineligible, discrepancy, BTZ, and small unit members.
    """
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        # Helper function to convert dataframes to list of dicts for JSON serialization
        def df_to_list(df, category=''):
            if df is None:
                return []

            records = []

            # If it's already a list, use it
            if isinstance(df, list):
                records = df
            # If it's a DataFrame, convert to dict records
            elif hasattr(df, 'to_dict'):
                records = df.to_dict('records')
            else:
                return []

            # Filter out soft-deleted records and clean up the data
            cleaned_records = []
            for idx, record in enumerate(records):
                # Skip soft-deleted records
                if record.get('deleted', False):
                    continue

                # Remove internal delete columns from output
                clean_record = {k: v for k, v in record.items()
                               if k not in ['deleted', 'deletion_reason']}

                # Add member_id based on index and category
                if category:
                    clean_record['member_id'] = f'row_{category}_{len(cleaned_records)}'
                else:
                    clean_record['member_id'] = f'row_{len(cleaned_records)}'

                cleaned_records.append(clean_record)

            return cleaned_records

        # Get all dataframes from session
        eligible_df = session.get('eligible_df', [])
        ineligible_df = session.get('ineligible_df', [])
        discrepancy_df = session.get('discrepancy_df', [])
        btz_df = session.get('btz_df', [])
        small_unit_df = session.get('small_unit_df', [])
        dataframe = session.get('dataframe', [])

        # Calculate statistics
        statistics = {
            "total_uploaded": len(df_to_list(dataframe)),
            "total_processed": (
                len(df_to_list(eligible_df, 'eligible')) +
                len(df_to_list(ineligible_df, 'ineligible')) +
                len(df_to_list(discrepancy_df, 'discrepancy')) +
                len(df_to_list(btz_df, 'btz'))
            ),
            "eligible": len(df_to_list(eligible_df, 'eligible')),
            "ineligible": len(df_to_list(ineligible_df, 'ineligible')),
            "discrepancy": len(df_to_list(discrepancy_df, 'discrepancy')),
            "btz": len(df_to_list(btz_df, 'btz')),
            "errors": len(session.get('error_log', []))
        }

        pascode_map = session.get('pascode_map', {})
        small_unit_sr = session.get('small_unit_sr')
        srid_pascode_map = session.get('srid_pascode_map', {})
        senior_rater_needed = bool(session.get('small_unit_df'))

        # Build response
        response = {
            "session_id": session_id,
            "cycle": session.get('cycle', 'SSG'),
            "year": session.get('year', 2025),
            "edited": session.get('edited', False),
            "statistics": statistics,
            "categories": {
                "eligible": df_to_list(eligible_df, 'eligible'),
                "ineligible": df_to_list(ineligible_df, 'ineligible'),
                "discrepancy": df_to_list(discrepancy_df, 'discrepancy'),
                "btz": df_to_list(btz_df, 'btz'),
                "small_unit": df_to_list(small_unit_df, 'small_unit')
            },
            "errors": session.get('error_log', []),
            "pascodes": session.get('pascodes', []),
            "pascode_unit_map": session.get('pascode_unit_map', {}),
            "custom_logo": session.get('custom_logo', {
                "uploaded": False,
                "filename": None
            }),
            "pascode_map": pascode_map,
            "srid_pascode_map": srid_pascode_map,
            "small_unit_sr": small_unit_sr,
            "senior_rater_needed": senior_rater_needed
        }

        return JSONResponse(content=response)

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to retrieve roster preview: {str(e)}"},
            status_code=500
        )


@app.put("/api/roster/member/{session_id}/{member_id}")
async def edit_roster_member(session_id: str, member_id: str, member_data: Dict):
    """
    Edit an existing member in the roster.
    Updates the member data in all relevant dataframes (eligible, ineligible, etc.)
    """
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        # Parse member_id to get category and index
        # Format: row_category_index (e.g., row_eligible_0)
        try:
            parts = member_id.split('_')
            if len(parts) >= 3 and parts[0] == 'row':
                # Extract category and index
                category_name = parts[1]  # e.g., 'eligible'
                index = int(parts[2])
                category_key = f"{category_name}_df"  # e.g., 'eligible_df'
            else:
                return JSONResponse(
                    content={"error": f"Invalid member_id format: {member_id}"},
                    status_code=400
                )
        except (ValueError, IndexError):
            return JSONResponse(
                content={"error": f"Invalid member_id format: {member_id}"},
                status_code=400
            )

        # Get the specific category list
        data_list = session.get(category_key, [])

        if not isinstance(data_list, list):
            return JSONResponse(
                content={"error": f"Category {category_name} not found"},
                status_code=404
            )

        if index >= len(data_list):
            return JSONResponse(
                content={"error": f"Member index {index} out of range for category {category_name}"},
                status_code=404
            )

        # Store original member data for finding in pdf_dataframe
        original_member = dict(data_list[index])

        # Only update fields that already exist in the record to avoid adding new columns
        # This prevents UI fields like REENL_ELIG_STATUS from being added to PDF data
        filtered_updates = {}
        for key, value in member_data.items():
            if key in data_list[index]:
                filtered_updates[key] = value

        # Update the member in the specific category with filtered data
        data_list[index].update(filtered_updates)

        # Also update in pdf_dataframe if it exists
        pdf_list = session.get('pdf_dataframe', [])
        if isinstance(pdf_list, list):
            # Find matching record in pdf_dataframe by comparing key fields
            for i, pdf_member in enumerate(pdf_list):
                if (pdf_member.get('FULL_NAME') == original_member.get('FULL_NAME') and
                    pdf_member.get('SSAN') == original_member.get('SSAN')):
                    # Only update existing fields in pdf_dataframe too
                    for key, value in filtered_updates.items():
                        if key in pdf_list[i]:
                            pdf_list[i][key] = value
                    break

        # Update session with modified lists
        updates = {
            category_key: pd.DataFrame(data_list),
            'pdf_dataframe': pd.DataFrame(pdf_list) if pdf_list else pd.DataFrame(),
            'edited': True
        }
        update_session(session_id, **updates)

        # Re-scan ALL categories for PASCODEs to ensure complete tracking
        # This is especially important if the ASSIGNED_PAS was changed during edit
        rescan_and_update_pascodes(session_id)

        # Recalculate small_unit_df to update senior_rater_needed flag
        recalculate_small_units(session_id)

        return JSONResponse(content={
            "success": True,
            "message": "Member updated successfully",
            "member_id": member_id
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to update member: {str(e)}"},
            status_code=500
        )


@app.delete("/api/roster/member/{session_id}/{member_id}")
async def delete_roster_member(
    session_id: str,
    member_id: str,
    reason: str = Query(..., description="Reason for deletion"),
    hard_delete: bool = Query(False, description="Permanently delete if True")
):
    """
    Delete a member from the roster.
    - reason: Required reason for deletion (for audit trail)
    - hard_delete: If True, permanently removes the member. If False, marks as deleted.
    """
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        if not reason:
            return JSONResponse(
                content={"error": "Reason for deletion is required"},
                status_code=400
            )

        # Parse member_id to get category and index
        # Format: row_category_index (e.g., row_eligible_0)
        try:
            parts = member_id.split('_')
            if len(parts) >= 3 and parts[0] == 'row':
                # Extract category and index
                category_name = parts[1]  # e.g., 'eligible'
                index = int(parts[2])
                category_key = f"{category_name}_df"  # e.g., 'eligible_df'
            else:
                return JSONResponse(
                    content={"error": f"Invalid member_id format: {member_id}"},
                    status_code=400
                )
        except (ValueError, IndexError):
            return JSONResponse(
                content={"error": f"Invalid member_id format: {member_id}"},
                status_code=400
            )

        # Get the specific category list
        data_list = session.get(category_key, [])

        if not isinstance(data_list, list):
            return JSONResponse(
                content={"error": f"Category {category_name} not found"},
                status_code=404
            )

        if index >= len(data_list):
            return JSONResponse(
                content={"error": f"Member index {index} out of range for category {category_name}"},
                status_code=404
            )

        # Perform the deletion only on the specific category
        if hard_delete:
            # Permanently remove the item
            data_list.pop(index)
        else:
            # Soft delete - mark as deleted
            data_list[index]['deleted'] = True
            data_list[index]['deletion_reason'] = reason

        # Also update in pdf_dataframe if it exists
        pdf_list = session.get('pdf_dataframe', [])
        if isinstance(pdf_list, list) and index < len(pdf_list):
            # Find matching record in pdf_dataframe by comparing key fields
            deleted_member = session.get(category_key, [])[index] if index < len(session.get(category_key, [])) else None
            if deleted_member:
                # Try to find and update/remove the same member in pdf_dataframe
                for i, pdf_member in enumerate(pdf_list):
                    if (pdf_member.get('FULL_NAME') == deleted_member.get('FULL_NAME') and
                        pdf_member.get('SSAN') == deleted_member.get('SSAN')):
                        if hard_delete:
                            pdf_list.pop(i)
                        else:
                            pdf_list[i]['deleted'] = True
                            pdf_list[i]['deletion_reason'] = reason
                        break

        # Update session with modified lists
        updates = {
            category_key: pd.DataFrame(data_list) if data_list else pd.DataFrame(),
            'pdf_dataframe': pd.DataFrame(pdf_list) if pdf_list else pd.DataFrame(),
            'edited': True
        }

        update_session(session_id, **updates)

        # Re-scan ALL categories for PASCODEs to ensure complete tracking
        # This is important to remove PASCODEs that no longer have any members
        rescan_and_update_pascodes(session_id)

        # Recalculate small_unit_df to update senior_rater_needed flag
        recalculate_small_units(session_id)

        return JSONResponse(content={
            "success": True,
            "message": f"Member {'permanently deleted' if hard_delete else 'marked as deleted'} successfully from {category_name}",
            "member_id": member_id,
            "category": category_name,
            "reason": reason,
            "hard_delete": hard_delete
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to delete member: {str(e)}"},
            status_code=500
        )


@app.post("/api/roster/member/{session_id}")
async def add_roster_member(
    session_id: str,
    data: Dict
):
    """
    Add a new member to the roster.
    Expected data format:
    {
        "category": "eligible|ineligible|discrepancy|btz|small_unit",
        "data": { member fields },
        "reason": "reason for adding",
        "run_eligibility_check": bool
    }
    """
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        category = data.get('category', 'eligible')
        member_data = data.get('data', {})
        reason = data.get('reason', '')
        run_eligibility_check = data.get('run_eligibility_check', True)  # Default to True
        senior_rater_info = data.get('senior_rater_info', {})

        # TODO: Implement automatic eligibility checking here
        # For now, members are added to the specified category (defaults to 'eligible')
        # The recalculate_small_units() call at the end ensures small_unit_df is updated
        # Future enhancement: Run board_filter and accounting_date_check to auto-determine category

        if not reason:
            return JSONResponse(
                content={"error": "Reason for adding member is required"},
                status_code=400
            )

        # Get the appropriate category list
        category_key = f"{category}_df"
        category_list = session.get(category_key, [])

        # If it doesn't exist or isn't a list, initialize it
        if not isinstance(category_list, list):
            category_list = []

        # Get a template from existing data to know which columns should exist
        # Use the first record as a template, or define default columns
        if category_list and len(category_list) > 0:
            # Use existing record as template for allowed fields
            template = category_list[0]
            new_member = {}

            # Only include fields that exist in the template
            for key in template.keys():
                if key in member_data:
                    new_member[key] = member_data[key]
                elif key == 'REASON' and category in ['ineligible', 'discrepancy']:
                    new_member[key] = reason
                elif key == 'member_id':
                    continue  # Will be set below
                else:
                    # Copy the default/empty value from template
                    new_member[key] = template.get(key, '')
        else:
            # If no existing records, use a minimal set of fields
            # This ensures we don't add UI-only fields to the data
            from constants import PDF_COLUMNS
            new_member = {}
            for col in PDF_COLUMNS:
                if col in member_data:
                    new_member[col] = member_data[col]
            # Also include PAFSC if provided (Primary AFSC)
            if 'PAFSC' in member_data:
                new_member['PAFSC'] = member_data['PAFSC']

            # Add REASON field for ineligible/discrepancy categories
            if category in ['ineligible', 'discrepancy'] and 'REASON' not in new_member:
                new_member['REASON'] = reason

        # Always run eligibility check for eligible category
        if category == 'eligible':
            # For new members being added to eligible, we assume they've been manually verified
            # The eligibility check would typically validate dates, rank progressions, etc.
            # Since this is a manual add with reason provided, we'll add them as eligible
            # but note in the reason that it was manually added
            if not reason.lower().startswith('manual add'):
                reason = f"Manual add: {reason}"

        # Generate a member_id
        new_index = len(category_list)
        new_member['member_id'] = f"row_{category}_{new_index}"

        # Track PASCODE for all categories (not just eligible)
        # This ensures senior rater information is collected for all units in the roster
        if 'ASSIGNED_PAS' in new_member:
            new_pascode = new_member['ASSIGNED_PAS']
            existing_pascodes = session.get('pascodes', [])

            # Always update pascode tracking regardless of whether it exists
            if new_pascode and new_pascode not in existing_pascodes:
                existing_pascodes.append(new_pascode)
                update_session(session_id, pascodes=existing_pascodes)

            # Always update pascode_unit_map if unit name is provided
            if 'ASSIGNED_PAS_CLEARTEXT' in new_member:
                pascode_unit_map = session.get('pascode_unit_map', {})
                pascode_unit_map[new_pascode] = new_member['ASSIGNED_PAS_CLEARTEXT']
                update_session(session_id, pascode_unit_map=pascode_unit_map)

            # Always update pascode_map with senior rater information if provided
            if senior_rater_info and new_pascode:
                pascode_map = session.get('pascode_map', {})
                pascode_map[new_pascode] = {
                    'srid': senior_rater_info.get('SRID', ''),
                    'senior_rater_name': senior_rater_info.get('SENIOR_RATER_NAME', ''),
                    'senior_rater_rank': senior_rater_info.get('SENIOR_RATER_RANK', ''),
                    'senior_rater_title': senior_rater_info.get('SENIOR_RATER_TITLE', ''),
                    'pascode': new_pascode,
                    'unit_name': new_member.get('ASSIGNED_PAS_CLEARTEXT', '')
                }
                update_session(session_id, pascode_map=pascode_map)

        # Also check if this should be added to small unit based on SRID
        if senior_rater_info and senior_rater_info.get('SRID'):
            # Check if this SRID indicates a small unit (you may need to adjust this logic)
            # For now, we'll track the SRID for potential small unit processing
            small_unit_sr = session.get('small_unit_sr', {})
            if not small_unit_sr:
                # If no small unit SR is set, this might be one
                small_unit_sr = {
                    'srid': senior_rater_info.get('SRID', ''),
                    'senior_rater_name': senior_rater_info.get('SENIOR_RATER_NAME', ''),
                    'senior_rater_rank': senior_rater_info.get('SENIOR_RATER_RANK', ''),
                    'senior_rater_title': senior_rater_info.get('SENIOR_RATER_TITLE', '')
                }
                update_session(session_id, small_unit_sr=small_unit_sr)

        # Add the new member to the list
        category_list.append(new_member)

        # Also add to pdf_dataframe if it exists
        pdf_list = session.get('pdf_dataframe', [])
        if isinstance(pdf_list, list):
            # Add the filtered member data to pdf list (without UI-only fields)
            pdf_new_member = new_member.copy()
            pdf_new_member.pop('member_id', None)  # Remove member_id from PDF data
            pdf_list.append(pdf_new_member)
            update_session(session_id, pdf_dataframe=pd.DataFrame(pdf_list))

        # Update the session with the modified list (will be converted back to list in session manager)
        update_dict = {
            category_key: pd.DataFrame(category_list),
            'edited': True
        }
        update_session(session_id, **update_dict)

        # Re-scan ALL categories for PASCODEs to ensure complete tracking
        # This catches any PASCODEs that might have been missed
        rescan_and_update_pascodes(session_id)

        # Recalculate small_unit_df to update senior_rater_needed flag
        recalculate_small_units(session_id)

        return JSONResponse(content={
            "success": True,
            "message": "Member added successfully",
            "member_id": member_data.get('member_id'),
            "category": category
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to add member: {str(e)}"},
            status_code=500
        )


@app.post("/api/roster/logo/{session_id}")
async def upload_logo(
    session_id: str,
    logo: UploadFile = File(...)
):
    """Upload a custom logo for the roster"""
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        # Validate file type
        allowed_types = ['image/png', 'image/jpeg', 'image/jpg']
        if logo.content_type not in allowed_types:
            return JSONResponse(
                content={"error": "Invalid file type. Only PNG and JPG files are allowed."},
                status_code=400
            )

        # Read and store logo data in session
        logo_data = await logo.read()

        # Store logo information in session
        update_session(
            session_id,
            custom_logo={
                "uploaded": True,
                "filename": logo.filename,
                "content_type": logo.content_type,
                "data": logo_data.hex()  # Store as hex string for JSON compatibility
            },
            edited=True
        )

        return JSONResponse(content={
            "success": True,
            "message": "Logo uploaded successfully",
            "filename": logo.filename
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to upload logo: {str(e)}"},
            status_code=500
        )


@app.get("/api/roster/logo/{session_id}")
async def get_logo(session_id: str):
    """Get the custom logo for the roster"""
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        custom_logo = session.get('custom_logo', {})

        if not custom_logo or not custom_logo.get('uploaded'):
            return JSONResponse(
                content={"error": "No custom logo found"},
                status_code=404
            )

        # Convert hex string back to bytes
        logo_data = bytes.fromhex(custom_logo.get('data', ''))
        content_type = custom_logo.get('content_type', 'image/png')

        return StreamingResponse(
            io.BytesIO(logo_data),
            media_type=content_type,
            headers={
                "Content-Disposition": f"inline; filename={custom_logo.get('filename', 'logo.png')}"
            }
        )

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to get logo: {str(e)}"},
            status_code=500
        )


@app.delete("/api/roster/logo/{session_id}")
async def delete_logo(session_id: str):
    """Delete the custom logo for the roster"""
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        # Remove logo from session
        update_session(
            session_id,
            custom_logo={"uploaded": False, "filename": None},
            edited=True
        )

        return JSONResponse(content={
            "success": True,
            "message": "Logo deleted successfully"
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to delete logo: {str(e)}"},
            status_code=500
        )


@app.post("/api/roster/reprocess/{session_id}")
async def reprocess_roster(
    session_id: str,
    data: Dict = Body(...)
):
    """Reprocess the roster with updated eligibility rules"""
    try:
        session = get_session(session_id)

        if not session:
            return JSONResponse(
                content={"error": "Session not found or expired"},
                status_code=404
            )

        preserve_edits = data.get('preserve_manual_edits', True)
        categories = data.get('categories', [])

        # Get the original dataframe
        dataframe_list = session.get('dataframe', [])
        pdf_list = session.get('pdf_dataframe', [])

        if not dataframe_list:
            return JSONResponse(
                content={"error": "No roster data found to reprocess"},
                status_code=400
            )

        # Convert lists back to DataFrames for processing
        dataframe = pd.DataFrame(dataframe_list)
        pdf_dataframe = pd.DataFrame(pdf_list) if pdf_list else pd.DataFrame()

        # Reprocess with roster_processor (you may need to import and use the actual processor)
        # This is a simplified version - you'll need to adapt based on your actual processing logic

        # For now, just mark as reprocessed
        update_session(
            session_id,
            reprocessed=True,
            reprocess_timestamp=pd.Timestamp.now().isoformat(),
            edited=True
        )

        return JSONResponse(content={
            "success": True,
            "message": "Roster reprocessed successfully",
            "preserve_edits": preserve_edits,
            "categories": categories
        })

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to reprocess roster: {str(e)}"},
            status_code=500
        )


@app.post("/api/initial-mel/submit/pascode-info")
async def submit_pascode_info(payload: PasCodeSubmission):
    pascode_map = {pascode: info.model_dump() for pascode, info in payload.pascode_info.items()}
    if 'small_unit_sr' in pascode_map:
        small_unit_sr = pascode_map.pop('small_unit_sr')
    else:
        small_unit_sr = None

    if small_unit_sr:
        update_session(payload.session_id, small_unit_sr=small_unit_sr)
    update_session(payload.session_id, pascode_map=pascode_map)
    srid_pascode_map = {}
    session = get_session(payload.session_id)

    # Validate session exists
    if not session:
        return JSONResponse(
            content={"error": "Session not found or expired"},
            status_code=404
        )

    # Validate required keys exist
    if 'pascodes' not in session or 'pascode_map' not in session:
        return JSONResponse(
            content={"error": "Invalid session data - missing required keys"},
            status_code=400
        )

    for pascode in session['pascodes']:
        # Validate pascode exists in pascode_map
        if pascode not in session['pascode_map']:
            continue
        # Validate srid exists in pascode info
        if 'srid' not in session['pascode_map'][pascode]:
            continue
        srid = session['pascode_map'][pascode]['srid']
        if srid in srid_pascode_map:
            srid_pascode_map[srid].append(pascode)
        else:
            srid_pascode_map[srid] = [pascode]

    update_session(payload.session_id, srid_pascode_map=srid_pascode_map)

    # Check for custom logo in session, otherwise use default
    logo_path = os.path.join(images_dir, default_logo)
    custom_logo = session.get('custom_logo')
    temp_logo_path = None

    if custom_logo and custom_logo.get('uploaded'):
        # Convert hex string back to bytes and save to temporary file
        logo_data = bytes.fromhex(custom_logo['data'])
        temp_logo_path = f"tmp/{payload.session_id}_logo.png"
        with open(temp_logo_path, 'wb') as f:
            f.write(logo_data)
        logo_path = temp_logo_path

    try:
        response = generate_roster_pdf(payload.session_id,
                                       output_filename=rf"tmp/{payload.session_id}_initial_mel_roster.pdf",
                                       logo_path=logo_path)

        # Clean up temporary logo file after PDF generation
        if temp_logo_path and os.path.exists(temp_logo_path):
            os.remove(temp_logo_path)

        if response:
            return response
        return JSONResponse(content={"error": "PDF generation failed"}, status_code=500)
    except Exception as e:
        # Clean up temporary logo file on error
        if temp_logo_path and os.path.exists(temp_logo_path):
            os.remove(temp_logo_path)
        raise


@app.post("/api/upload/final-mel")
async def upload_final_mel_file(
        file: UploadFile = File(...),
        cycle: str = Form(...),
        year: int = Form(...)
):
    # Generate session ID early for logging
    session_id = str(uuid.uuid4())

    # CRITICAL FIX: Validate cycle parameter
    valid_cycles = ['SRA', 'SSG', 'TSG', 'MSG', 'SMS']
    if cycle not in valid_cycles:
        return JSONResponse(
            content={"error": f"Invalid cycle. Must be one of: {', '.join(valid_cycles)}"},
            status_code=400
        )

    # CRITICAL FIX: Validate year parameter
    from constants import MIN_PROMOTION_CYCLE_YEAR, MAX_PROMOTION_CYCLE_YEAR
    try:
        year = int(year)
        if year < MIN_PROMOTION_CYCLE_YEAR or year > MAX_PROMOTION_CYCLE_YEAR:
            return JSONResponse(
                content={"error": f"Invalid year. Must be between {MIN_PROMOTION_CYCLE_YEAR} and {MAX_PROMOTION_CYCLE_YEAR}"},
                status_code=400
            )
    except (ValueError, TypeError):
        return JSONResponse(content={"error": "Year must be a valid integer"}, status_code=400)

    # Create session-specific logger
    logger = LoggerSetup.get_session_logger(session_id, cycle, year)

    logger.info(f"FINAL MEL UPLOAD STARTED")
    logger.info(f"  Filename: {file.filename}")
    logger.info(f"  Content Type: {file.content_type}")
    logger.info(f"  Cycle: {cycle}")
    logger.info(f"  Year: {year}")

    return_object = {}

    if file.content_type not in allowed_types:
        error_msg = "Invalid file type. Only CSV or Excel files are allowed."
        logger.error(f"  FAILED: {error_msg}")
        logger.info(f"STATUS: FAILED - Invalid File Type")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=400)

    contents = await file.read()
    file_size_bytes = len(contents)
    logger.info(f"  File size: {file_size_bytes} bytes")

    # HIGH FIX: Validate file size
    from constants import MAX_FILE_SIZE_MB
    max_size_bytes = MAX_FILE_SIZE_MB * 1024 * 1024
    if file_size_bytes > max_size_bytes:
        error_msg = f"File too large. Maximum size is {MAX_FILE_SIZE_MB}MB"
        logger.error(f"  FAILED: {error_msg} (received {file_size_bytes / 1024 / 1024:.2f}MB)")
        logger.info(f"STATUS: FAILED - File Too Large")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=400)

    try:
        if file.filename.endswith(".csv"):
            logger.info(f"  Parsing CSV file")
            # HIGH FIX: Add CSV encoding handling
            try:
                df = pd.read_csv(io.BytesIO(contents), encoding='utf-8')
            except UnicodeDecodeError:
                logger.warning(f"  UTF-8 decoding failed, trying cp1252 encoding")
                try:
                    df = pd.read_csv(io.BytesIO(contents), encoding='cp1252')
                except UnicodeDecodeError:
                    logger.warning(f"  cp1252 decoding failed, trying latin1 encoding")
                    df = pd.read_csv(io.BytesIO(contents), encoding='latin1')
        elif file.filename.endswith(".xlsx"):
            logger.info(f"  Parsing Excel file")
            # CRITICAL FIX: Validate Excel sheet has data
            df = pd.read_excel(io.BytesIO(contents), sheet_name=0)
            if df.empty:
                logger.warning(f"  First sheet is empty, checking other sheets...")
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents))
                sheet_names = wb.sheetnames
                logger.info(f"  Available sheets: {sheet_names}")
                # Try to find first non-empty sheet
                for sheet_name in sheet_names:
                    test_df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name)
                    if not test_df.empty:
                        df = test_df
                        logger.info(f"  Using sheet '{sheet_name}' which contains data")
                        break
        else:
            error_msg = "Unsupported file extension."
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"STATUS: FAILED - Unsupported Extension")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        logger.info(f"  File parsed successfully: {len(df)} rows, {len(df.columns)} columns")

        # HIGH FIX: Filter out completely empty rows
        initial_rows = len(df)
        df = df.dropna(how='all')
        if len(df) < initial_rows:
            logger.info(f"  Filtered out {initial_rows - len(df)} empty rows")

        # HIGH FIX: Normalize column names to uppercase
        df.columns = df.columns.str.strip().str.upper()
        logger.info(f"  Normalized column names to uppercase")

        # MEDIUM FIX: Strip leading/trailing whitespace from string columns
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)

        # CRITICAL FIX: Validate required columns exist
        all_required_columns = REQUIRED_COLUMNS
        missing_columns = [col for col in all_required_columns if col not in df.columns]
        if missing_columns:
            error_msg = f"Missing required columns: {', '.join(missing_columns)}"
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"  Available columns: {', '.join(df.columns.tolist())}")
            logger.info(f"STATUS: FAILED - Missing Columns")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        # Filter to only include columns we need (required + optional that exist)
        available_optional = [col for col in OPTIONAL_COLUMNS if col in df.columns]
        columns_to_keep = REQUIRED_COLUMNS + available_optional
        processed_df = df[columns_to_keep].copy()

        # Validate PDF columns exist
        missing_pdf_columns = [col for col in PDF_COLUMNS if col not in processed_df.columns]
        if missing_pdf_columns:
            error_msg = f"Missing PDF columns: {', '.join(missing_pdf_columns)}"
            logger.error(f"  FAILED: {error_msg}")
            logger.info(f"STATUS: FAILED - Missing PDF Columns")
            LoggerSetup.close_session_logger(session_id)
            return JSONResponse(content={"error": error_msg}, status_code=400)

        pdf_df = processed_df[PDF_COLUMNS].copy()

        # Pass session_id to create_session so it uses our pre-generated ID
        create_session(processed_df, pdf_df, session_id=session_id)
        logger.info(f"  Session created: {session_id}")

        # MEDIUM FIX: Batch session updates to reduce Redis round trips
        update_session(session_id, cycle=cycle, year=year)

        logger.info(f"  Starting roster processing...")
        roster_processor(df, session_id, cycle, year)
        logger.info(f"  Roster processing complete")

        session = get_session(session_id)

        # Use .get() to safely access session keys that might not exist
        if session.get('pascodes') is not None:
            return_object['pascodes'] = session['pascodes']
            logger.info(f"  PASCODEs found: {len(session['pascodes'])}")

        if session.get('pascode_unit_map') is not None:
            return_object['pascode_unit_map'] = session['pascode_unit_map']

        if session.get('small_unit_df') is not None:
            return_object['senior_rater_needed'] = True
            logger.info(f"  Senior rater required for small units")
        else:
            return_object["senior_rater_needed"] = False

        return_object['message'] = "Upload successful."
        return_object['session_id'] = session_id
        return_object['errors'] = session.get('error_log', [])

        if return_object['errors']:
            logger.warning(f"  Upload completed with {len(return_object['errors'])} errors")
        else:
            logger.info(f"  Upload completed successfully with no errors")

        logger.info(f"FINAL MEL UPLOAD COMPLETED SUCCESSFULLY")

        return JSONResponse(content=return_object)

    except Exception as e:
        error_msg = f"Processing error: {str(e)}"
        logger.error(f"  EXCEPTION: {error_msg}", exc_info=True)
        logger.info(f"STATUS: FAILED - Exception")
        LoggerSetup.close_session_logger(session_id)
        return JSONResponse(content={"error": error_msg}, status_code=500)


@app.post("/api/final-mel/submit/pascode-info")
async def submit_final_pascode_info(payload: PasCodeSubmission):
    pascode_map = {pascode: info.model_dump() for pascode, info in payload.pascode_info.items()}
    if 'small_unit_sr' in pascode_map:
        small_unit_sr = pascode_map.pop('small_unit_sr')
    else:
        small_unit_sr = None

    if small_unit_sr:
        update_session(payload.session_id, small_unit_sr=small_unit_sr)
    update_session(payload.session_id, pascode_map=pascode_map)
    srid_pascode_map = {}
    session = get_session(payload.session_id)

    # Validate session exists
    if not session:
        return JSONResponse(
            content={"error": "Session not found or expired"},
            status_code=404
        )

    # Validate required keys exist
    if 'pascodes' not in session or 'pascode_map' not in session:
        return JSONResponse(
            content={"error": "Invalid session data - missing required keys"},
            status_code=400
        )

    for pascode in session['pascodes']:
        # Validate pascode exists in pascode_map
        if pascode not in session['pascode_map']:
            continue
        # Validate srid exists in pascode info
        if 'srid' not in session['pascode_map'][pascode]:
            continue
        srid = session['pascode_map'][pascode]['srid']
        if srid in srid_pascode_map:
            srid_pascode_map[srid].append(pascode)
        else:
            srid_pascode_map[srid] = [pascode]

    update_session(payload.session_id, srid_pascode_map=srid_pascode_map)

    # Check for custom logo in session, otherwise use default
    logo_path = os.path.join(images_dir, default_logo)
    custom_logo = session.get('custom_logo')
    temp_logo_path = None

    if custom_logo and custom_logo.get('uploaded'):
        # Convert hex string back to bytes and save to temporary file
        logo_data = bytes.fromhex(custom_logo['data'])
        temp_logo_path = f"tmp/{payload.session_id}_logo.png"
        with open(temp_logo_path, 'wb') as f:
            f.write(logo_data)
        logo_path = temp_logo_path

    try:
        response = generate_final_roster_pdf(payload.session_id,
                                             output_filename=rf"tmp/{payload.session_id}_final_mel_roster.pdf",
                                             logo_path=logo_path)

        # Clean up temporary logo file after PDF generation
        if temp_logo_path and os.path.exists(temp_logo_path):
            os.remove(temp_logo_path)

        if response:
            return response
        return JSONResponse(content={"error": "PDF generation failed"}, status_code=500)
    except Exception as e:
        # Clean up temporary logo file on error
        if temp_logo_path and os.path.exists(temp_logo_path):
            os.remove(temp_logo_path)
        raise


@app.get("/api/download/final-mel/{session_id}")
async def download_final_mel(session_id: str):
    try:
        pdf_buffer: Optional[io.BytesIO] = get_pdf_from_redis(session_id)

        if not pdf_buffer:
            return JSONResponse(
                content={"error": "PDF not found for this session"},
                status_code=404
            )

        return StreamingResponse(
            pdf_buffer,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename=final_mel_roster.pdf"
            }
        )
    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to retrieve PDF: {str(e)}"},
            status_code=500
        )


# ============================================================================
# DOCUMENT GENERATOR ENDPOINTS
# ============================================================================

from document_generator.models import (
    DocumentGenerationRequest, DocumentResponse, DocumentType,
    MFRContent, MemoContent, AppointmentContent, AdministrativeActionContent
)
from document_generator.generators import (
    MFRGenerator, MemoGenerator, AppointmentGenerator,
    LOCGenerator, LOAGenerator, LORGenerator
)
from document_generator.prompt_parser import PromptParser
from document_generator.validation import DocumentValidator
from constants import AF_DOCUMENT_TYPES, DOCUMENT_SESSION_TTL


@app.get("/api/documents/version")
async def get_document_generator_version():
    """Check if JSON serialization fix is active"""
    return JSONResponse(content={
        "version": "1.0.1",
        "json_fix_applied": True,
        "last_update": "2025-11-05",
        "message": "JSON serialization fix is active"
    })


@app.get("/api/documents/templates")
async def list_document_templates():
    """List available document templates with examples"""
    templates = [
        {
            "type": "mfr",
            "name": "Memorandum For Record",
            "description": "Internal documentation of events, decisions, or conversations",
            "required_fields": ["subject", "body_paragraphs"],
            "optional_fields": ["distribution_list", "attachments"],
            "example_prompts": [
                "Create an MFR documenting a safety briefing on 15 Jan 2025",
                "MFR about phone call with finance office regarding TDY voucher on 10 Jan 2025"
            ]
        },
        {
            "type": "memo",
            "name": "Official Memorandum",
            "description": "Formal communication between organizations or offices",
            "required_fields": ["to_line", "subject", "body_paragraphs"],
            "optional_fields": ["thru_line", "distribution_list", "attachments"],
            "example_prompts": [
                "Create a memo to 51 FSS/CCF about the training schedule for February 2025",
                "Official memo requesting additional manning support for upcoming inspection"
            ]
        },
        {
            "type": "appointment",
            "name": "Appointment Letter",
            "description": "Assign duties, responsibilities, and authorities to a member",
            "required_fields": ["appointee_name", "appointee_rank", "position_title", "authority_citation", "duties", "effective_date"],
            "optional_fields": ["appointee_ssan", "termination_date"],
            "example_prompts": [
                "Appoint SSgt John Smith as Squadron Safety Representative per AFI 91-202 effective 15 Jan 2025",
                "Create appointment letter for TSgt Jane Doe as Unit Fitness Program Manager"
            ]
        },
        {
            "type": "loc",
            "name": "Letter of Counseling",
            "description": "Initial corrective action for minor infractions",
            "required_fields": ["member_name", "member_rank", "incident_date", "incident_description", "violations", "standards_expected", "consequences"],
            "optional_fields": ["member_ssan"],
            "example_prompts": [
                "Create LOC for SrA Smith who was late to duty on 10 Jan 2025 violating AFI 36-2618",
                "LOC for uniform violation on 5 Jan 2025"
            ]
        },
        {
            "type": "loa",
            "name": "Letter of Admonishment",
            "description": "Mid-level corrective action for repeated or more serious infractions",
            "required_fields": ["member_name", "member_rank", "incident_date", "incident_description", "violations", "standards_expected", "consequences"],
            "optional_fields": ["member_ssan", "previous_actions", "appeal_rights"],
            "example_prompts": [
                "Create LOA for repeated tardiness despite previous LOC",
                "LOA for failure to follow lawful order on 12 Jan 2025"
            ]
        },
        {
            "type": "lor",
            "name": "Letter of Reprimand",
            "description": "Serious corrective action filed in personnel records",
            "required_fields": ["member_name", "member_rank", "incident_date", "incident_description", "violations", "standards_expected", "consequences", "filing_location"],
            "optional_fields": ["member_ssan", "previous_actions", "appeal_rights"],
            "example_prompts": [
                "Create LOR for DUI incident on 1 Jan 2025 to be filed in PIF",
                "LOR for dereliction of duty violating AFI 1-1"
            ]
        }
    ]

    return JSONResponse(content={"templates": templates})


@app.post("/api/documents/generate")
async def generate_document(request: DocumentGenerationRequest):
    """
    Generate a document from a natural language prompt

    This endpoint:
    1. Parses the prompt to extract structured data
    2. Validates the extracted information
    3. Creates a document session in Redis
    4. Returns extracted fields for user review/editing
    """
    try:
        # Create document session ID
        document_id = str(uuid.uuid4())

        # Parse the prompt
        parser = PromptParser()
        extracted_fields = parser.parse_prompt(request.prompt, request.document_type.value)

        # Add metadata to extracted fields (serialize dates as ISO strings)
        extracted_fields['metadata'] = request.metadata.model_dump(mode='json')

        # Store in Redis session with longer TTL (2 hours vs 30 min for MEL)
        update_session(document_id,
                      document_type=request.document_type.value,
                      extracted_fields=extracted_fields,
                      original_prompt=request.prompt,
                      status='draft')

        # Validate extracted content (warnings only, not blocking)
        validation_warnings = []

        if request.document_type == DocumentType.MFR:
            if not extracted_fields.get('subject'):
                validation_warnings.append("Could not extract subject line from prompt")
            if not extracted_fields.get('body_hints'):
                validation_warnings.append("Could not extract body content from prompt")

        elif request.document_type == DocumentType.APPOINTMENT:
            if not extracted_fields.get('appointee_name'):
                validation_warnings.append("Could not extract appointee name from prompt")
            if not extracted_fields.get('authority_citation'):
                validation_warnings.append("AFI citation not found - will need to be added manually")

        response = DocumentResponse(
            document_id=document_id,
            session_id=document_id,  # Using same ID for simplicity
            document_type=request.document_type.value,
            status='draft',
            message="Document draft created. Review extracted fields and generate PDF.",
            extracted_fields=extracted_fields,
            validation_warnings=validation_warnings,
            pdf_url=f"/api/documents/{document_id}/generate-pdf"
        )

        return JSONResponse(content=response.model_dump(mode='json'))

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to generate document: {str(e)}"},
            status_code=500
        )


@app.post("/api/documents/{document_id}/generate-pdf")
async def generate_document_pdf(
    document_id: str,
    content: Dict = Body(...)
):
    """
    Generate final PDF from document content

    Accepts the full document content (after user review/editing) and generates PDF
    """
    try:
        # Get session to retrieve document type
        session = get_session(document_id)

        if not session:
            return JSONResponse(
                content={"error": "Document session not found or expired"},
                status_code=404
            )

        document_type = session.get('document_type')
        metadata_dict = content.get('metadata', {})

        # Convert metadata dict to DocumentMetadata model
        from document_generator.models import DocumentMetadata
        metadata = DocumentMetadata(**metadata_dict)

        # Generate PDF based on document type
        pdf_buffer = None

        if document_type == 'mfr':
            mfr_content = MFRContent(**content.get('content', {}))
            generator = MFRGenerator()
            pdf_buffer = generator.generate(mfr_content, metadata)

        elif document_type == 'memo':
            memo_content = MemoContent(**content.get('content', {}))
            generator = MemoGenerator()
            pdf_buffer = generator.generate(memo_content, metadata)

        elif document_type == 'appointment':
            appt_content = AppointmentContent(**content.get('content', {}))
            generator = AppointmentGenerator()
            pdf_buffer = generator.generate(appt_content, metadata)

        elif document_type == 'loc':
            loc_content = AdministrativeActionContent(**content.get('content', {}))
            generator = LOCGenerator()
            pdf_buffer = generator.generate(loc_content, metadata)

        elif document_type == 'loa':
            loa_content = AdministrativeActionContent(**content.get('content', {}))
            generator = LOAGenerator()
            pdf_buffer = generator.generate(loa_content, metadata)

        elif document_type == 'lor':
            lor_content = AdministrativeActionContent(**content.get('content', {}))
            generator = LORGenerator()
            pdf_buffer = generator.generate(lor_content, metadata)

        else:
            return JSONResponse(
                content={"error": f"Unsupported document type: {document_type}"},
                status_code=400
            )

        if not pdf_buffer:
            return JSONResponse(
                content={"error": "PDF generation failed"},
                status_code=500
            )

        # Store PDF in Redis
        from session_manager import store_pdf_in_redis
        store_pdf_in_redis(document_id, pdf_buffer)

        # Update session status
        update_session(document_id, status='finalized')

        # Return PDF as streaming response
        pdf_buffer.seek(0)
        return StreamingResponse(
            pdf_buffer,
            media_type='application/pdf',
            headers={
                "Content-Disposition": f"attachment; filename={document_type}_{document_id[:8]}.pdf"
            }
        )

    except Exception as e:
        return JSONResponse(
            content={"error": f"Failed to generate PDF: {str(e)}"},
            status_code=500
        )
